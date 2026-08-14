#!/usr/bin/env python3
"""
Module d'export Excel pour NetPing Monitor
Génère des rapports Excel détaillés des SpeedTests
"""

import os
from datetime import datetime
from typing import Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl non installé. Installez avec: pip install openpyxl")


class SpeedTestExcelExporter:
    """Exportateur de rapports Excel pour SpeedTest"""
    
    # Styles
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    
    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    SUMMARY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, output_dir: str = "reports/speedtests"):
        """
        Initialise l'exportateur
        
        Args:
            output_dir: Répertoire de sortie
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def generate_daily_report(self, results: List[Dict], date: str = None, 
                              include_charts: bool = True) -> Optional[str]:
        """
        Génère un rapport Excel journalier
        
        Args:
            results: Liste des résultats du jour
            date: Date du rapport (optionnel)
            include_charts: Inclure les graphiques
            
        Returns:
            Chemin du fichier créé ou None
        """
        if not EXCEL_AVAILABLE:
            print("Excel export non disponible - openpyxl requis")
            return None
        
        if not results:
            print("Aucun résultat à exporter")
            return None
        
        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')
        
        # Créer le workbook
        wb = openpyxl.Workbook()
        
        # === Feuille 1: Résumé ===
        ws_summary = wb.active
        ws_summary.title = "Résumé"
        self._create_summary_sheet(ws_summary, results, date)
        
        # === Feuille 2: Détails ===
        ws_details = wb.create_sheet("Détails")
        self._create_details_sheet(ws_details, results)
        
        # === Feuille 3: Graphiques ===
        if include_charts and len(results) > 1:
            ws_charts = wb.create_sheet("Graphiques")
            self._create_charts_sheet(ws_charts, results)
        
        # Sauvegarder
        filename = f"speedtest_report_{date}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            wb.save(filepath)
            return filepath
        except Exception as e:
            print(f"Erreur lors de la sauvegarde: {e}")
            return None
    
    def _create_summary_sheet(self, ws, results: List[Dict], date: str):
        """Crée la feuille de résumé"""
        # Titre
        ws['A1'] = f"RAPPORT SPEEDTEST - {date}"
        ws['A1'].font = Font(bold=True, size=16, color="4472C4")
        ws.merge_cells('A1:D1')
        
        # Statistiques globales
        ws['A3'] = "STATISTIQUES GLOBALES"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].fill = self.HEADER_FILL
        ws.merge_cells('A3:D3')
        
        successful = [r for r in results if r.get('success')]
        
        stats = [
            ("Nombre total de tests", len(results)),
            ("Tests réussis", len(successful)),
            ("Tests échoués", len(results) - len(successful)),
            ("Taux de réussite", f"{len(successful)/len(results)*100:.1f}%" if results else "N/A"),
        ]
        
        row = 4
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Statistiques de performance
        if successful:
            pings = [r['ping'] for r in successful if r.get('ping') is not None]
            downloads = [r['download'] for r in successful if r.get('download') is not None]
            uploads = [r['upload'] for r in successful if r.get('upload') is not None]
            
            ws['A9'] = "PERFORMANCES RÉSEAU"
            ws['A9'].font = Font(bold=True, size=12)
            ws['A9'].fill = self.HEADER_FILL
            ws.merge_cells('A9:D9')
            
            perf_stats = []
            
            if pings:
                perf_stats.extend([
                    ("Ping moyen", f"{sum(pings)/len(pings):.1f} ms"),
                    ("Ping min", f"{min(pings):.1f} ms"),
                    ("Ping max", f"{max(pings):.1f} ms"),
                ])
            
            if downloads:
                perf_stats.extend([
                    ("Download moyen", f"{sum(downloads)/len(downloads):.2f} Mbps"),
                    ("Meilleur download", f"{max(downloads):.2f} Mbps"),
                    ("Plus faible download", f"{min(downloads):.2f} Mbps"),
                ])
                
                # Trouver les heures
                for r in successful:
                    if r.get('download') == max(downloads):
                        perf_stats.append(("Heure meilleur download", r.get('time', 'N/A')))
                    if r.get('download') == min(downloads):
                        perf_stats.append(("Heure plus faible download", r.get('time', 'N/A')))
            
            if uploads:
                perf_stats.extend([
                    ("Upload moyen", f"{sum(uploads)/len(uploads):.2f} Mbps"),
                    ("Meilleur upload", f"{max(uploads):.2f} Mbps"),
                    ("Plus faible upload", f"{min(uploads):.2f} Mbps"),
                ])
            
            row = 10
            for label, value in perf_stats:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Informations IP
            ws[f'A{row+1}'] = "INFORMATIONS RÉSEAU"
            ws[f'A{row+1}'].font = Font(bold=True, size=12)
            ws[f'A{row+1}'].fill = self.HEADER_FILL
            ws.merge_cells(f'A{row+1}:D{row+1}')
            
            unique_ips = list(set(r.get('public_ip') for r in successful if r.get('public_ip')))
            
            row += 2
            ws[f'A{row}'] = "IP publique(s)"
            ws[f'B{row}'] = ', '.join(unique_ips) if unique_ips else 'N/A'
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = "ISP"
            ws[f'B{row}'] = successful[0].get('isp', 'N/A')
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = "Localisation"
            location = f"{successful[0].get('city', '')}, {successful[0].get('country', '')}"
            ws[f'B{row}'] = location if location.strip(',') else 'N/A'
            ws[f'A{row}'].font = Font(bold=True)
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def _create_details_sheet(self, ws, results: List[Dict]):
        """Crée la feuille de détails"""
        # En-têtes
        headers = [
            "Date", "Heure", "IP Publique", "ISP", "Ville", "Pays",
            "Ping (ms)", "Download (Mbps)", "Upload (Mbps)", 
            "Jitter (ms)", "Perte paquets (%)", "Serveur",
            "Statut", "Observations"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER
        
        # Données
        for row_idx, result in enumerate(results, 2):
            data = [
                result.get('date', ''),
                result.get('time', ''),
                result.get('public_ip', ''),
                result.get('isp', ''),
                result.get('city', ''),
                result.get('country', ''),
                result.get('ping', '') if result.get('ping') is not None else '',
                round(result.get('download', 0), 2) if result.get('download') is not None else '',
                round(result.get('upload', 0), 2) if result.get('upload') is not None else '',
                result.get('jitter', '') if result.get('jitter') is not None else '',
                result.get('packet_loss', '') if result.get('packet_loss') is not None else '',
                result.get('server_name', ''),
                result.get('status', ''),
                result.get('status_message', result.get('error', ''))
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Colorer selon le statut
                status = result.get('status', '')
                if col == 13:  # Colonne Statut
                    if status == 'good':
                        cell.fill = self.SUCCESS_FILL
                    elif status == 'warning':
                        cell.fill = self.WARNING_FILL
                    elif status in ('critical', 'error'):
                        cell.fill = self.ERROR_FILL
        
        # Ajuster les largeurs
        column_widths = [12, 10, 15, 15, 12, 10, 12, 15, 15, 12, 15, 20, 10, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_charts_sheet(self, ws, results: List[Dict]):
        """Crée la feuille avec les graphiques"""
        # Titre
        ws['A1'] = "GRAPHIQUES D'ÉVOLUTION"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Préparer les données pour les graphiques
        successful = [r for r in results if r.get('success')]
        
        if len(successful) < 2:
            ws['A3'] = "Pas assez de données pour générer des graphiques"
            return
        
        # Données pour graphique Ping
        ws['A3'] = "Heure"
        ws['B3'] = "Ping (ms)"
        ws['C3'] = "Download (Mbps)"
        ws['D3'] = "Upload (Mbps)"
        
        for i, r in enumerate(successful, 4):
            ws[f'A{i}'] = r.get('time', '')
            ws[f'B{i}'] = r.get('ping', 0)
            ws[f'C{i}'] = r.get('download', 0)
            ws[f'D{i}'] = r.get('upload', 0)
        
        data_range = len(successful) + 3
        
        # Graphique Ping
        ping_chart = LineChart()
        ping_chart.title = "Évolution du Ping"
        ping_chart.y_axis.title = "ms"
        ping_chart.x_axis.title = "Heure"
        ping_chart.style = 10
        ping_chart.width = 15
        ping_chart.height = 8
        
        ping_data = Reference(ws, min_col=2, min_row=3, max_col=2, max_row=data_range)
        ping_cats = Reference(ws, min_col=1, min_row=4, max_row=data_range)
        ping_chart.add_data(ping_data, titles_from_data=True)
        ping_chart.set_categories(ping_cats)
        
        ws.add_chart(ping_chart, "F3")
        
        # Graphique Débit
        speed_chart = LineChart()
        speed_chart.title = "Évolution du Débit"
        speed_chart.y_axis.title = "Mbps"
        speed_chart.x_axis.title = "Heure"
        speed_chart.style = 10
        speed_chart.width = 15
        speed_chart.height = 8
        
        speed_data = Reference(ws, min_col=3, min_row=3, max_col=4, max_row=data_range)
        speed_chart.add_data(speed_data, titles_from_data=True)
        speed_chart.set_categories(ping_cats)
        
        ws.add_chart(speed_chart, "F20")
    
    def generate_custom_report(self, results: List[Dict], title: str = "Rapport SpeedTest",
                               start_date: str = None, end_date: str = None) -> Optional[str]:
        """
        Génère un rapport personnalisé
        
        Args:
            results: Liste des résultats
            title: Titre du rapport
            start_date: Date de début
            end_date: Date de fin
            
        Returns:
            Chemin du fichier créé
        """
        if not EXCEL_AVAILABLE or not results:
            return None
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport"
        
        # Titre
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16, color="4472C4")
        ws.merge_cells('A1:G1')
        
        if start_date and end_date:
            ws['A2'] = f"Période: {start_date} - {end_date}"
            ws['A2'].font = Font(italic=True)
        
        # Créer les détails
        self._create_details_sheet(ws, results)
        
        # Sauvegarder
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"speedtest_custom_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            wb.save(filepath)
            return filepath
        except Exception as e:
            print(f"Erreur: {e}")
            return None


# Test du module
if __name__ == "__main__":
    print("=" * 60)
    print("TEST DU MODULE EXCEL EXPORTER")
    print("=" * 60)
    
    if not EXCEL_AVAILABLE:
        print("\n❌ openpyxl non installé")
        print("   Installez avec: pip install openpyxl")
    else:
        exporter = SpeedTestExcelExporter("test_reports/speedtests")
        
        # Données de test
        test_results = [
            {
                'date': '2026-06-22',
                'time': '08:50:00',
                'success': True,
                'public_ip': '41.243.13.114',
                'isp': 'Orange',
                'city': 'Paris',
                'country': 'France',
                'ping': 25.5,
                'download': 150.2,
                'upload': 45.8,
                'jitter': 3.2,
                'packet_loss': 0.0,
                'server_name': 'Paris Server',
                'status': 'good',
                'status_message': 'Connexion stable'
            },
            {
                'date': '2026-06-22',
                'time': '12:00:00',
                'success': True,
                'public_ip': '41.243.13.114',
                'isp': 'Orange',
                'city': 'Paris',
                'country': 'France',
                'ping': 35.2,
                'download': 95.5,
                'upload': 42.1,
                'jitter': 5.1,
                'packet_loss': 0.0,
                'server_name': 'Paris Server',
                'status': 'warning',
                'status_message': 'Download faible'
            },
            {
                'date': '2026-06-22',
                'time': '17:25:00',
                'success': True,
                'public_ip': '41.243.13.114',
                'isp': 'Orange',
                'city': 'Paris',
                'country': 'France',
                'ping': 28.3,
                'download': 120.8,
                'upload': 44.5,
                'jitter': 2.8,
                'packet_loss': 0.0,
                'server_name': 'Paris Server',
                'status': 'good',
                'status_message': 'Connexion stable'
            },
            {
                'date': '2026-06-22',
                'time': '22:00:00',
                'success': False,
                'public_ip': '41.243.13.114',
                'isp': 'Orange',
                'city': 'Paris',
                'country': 'France',
                'ping': None,
                'download': None,
                'upload': None,
                'jitter': None,
                'packet_loss': None,
                'server_name': None,
                'status': 'error',
                'status_message': 'Timeout'
            }
        ]
        
        print("\nGénération du rapport Excel...")
        filepath = exporter.generate_daily_report(test_results, '2026-06-22')
        
        if filepath:
            print(f"✅ Rapport créé: {filepath}")
        else:
            print("❌ Erreur lors de la création")
        
        # Nettoyer
        import shutil
        if os.path.exists("test_reports"):
            shutil.rmtree("test_reports")
        
        print("\nTest terminé.")
